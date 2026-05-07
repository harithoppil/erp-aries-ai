"""Document Output MCP Server — generate proposal PDFs, quotes, and internal summaries.

Uses reportlab for PDF generation and openpyxl for Excel/CSV quote files.
"""

# PORTED — This MCP server is now registered in Next.js src/lib/mcp-gateway.ts
# Tool handlers run as async functions calling Prisma or Python microservice.

import csv
import io
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path

from mcp.server.fastmcp import FastMCP

doc_output_mcp = FastMCP(
    "Document Output MCP",
    instructions="Generate proposal PDFs, quote spreadsheets, and internal summary documents"
)

# Ensure output directories exist
for _dir in ("media/proposals", "media/quotes", "media/summaries"):
    Path(_dir).mkdir(parents=True, exist_ok=True)


@doc_output_mcp.tool()
async def generate_proposal_pdf(
    enquiry_id: str,
    content: str,
    client_name: str,
    pricing_data: str = "",
    output_format: str = "pdf",
) -> str:
    """Generate a professional proposal PDF using reportlab.

    Args:
        enquiry_id: UUID of the enquiry.
        content: Proposal body text (markdown or plain).
        client_name: Client company name.
        pricing_data: Optional JSON string with line items.
        output_format: "pdf" or "markdown".

    Returns:
        Path to the generated file.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            ListFlowable, ListItem
        )
    except ImportError:
        # Fallback to markdown if reportlab not installed
        output_format = "markdown"

    safe_client = client_name.replace(" ", "_").replace("/", "_")[:50]
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    if output_format == "pdf":
        file_name = f"proposal_{safe_client}_{enquiry_id[:8]}_{timestamp}.pdf"
        file_path = Path("media/proposals") / file_name

        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )
        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1a365d"),
            spaceAfter=30,
        )
        story.append(Paragraph("PROPOSAL", title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Meta table
        meta_data = [
            ["Client:", client_name],
            ["Enquiry ID:", enquiry_id],
            ["Date:", datetime.now(timezone.utc).strftime("%Y-%m-%d")],
            ["Prepared by:", "Aries AI Presales Consultant"],
        ]
        meta_table = Table(meta_data, colWidths=[2 * inch, 4 * inch])
        meta_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e2e8f0")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
            ("ALIGN", (0, 0), (0, -1), "RIGHT"),
            ("ALIGN", (1, 0), (1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.3 * inch))

        # Body content
        body_style = ParagraphStyle(
            "Body",
            parent=styles["BodyText"],
            fontSize=11,
            leading=16,
            spaceAfter=12,
        )
        for paragraph in content.split("\n\n"):
            if paragraph.strip():
                story.append(Paragraph(paragraph.strip().replace("\n", "<br/>"), body_style))

        # Pricing table
        if pricing_data:
            story.append(Spacer(1, 0.3 * inch))
            story.append(Paragraph("Pricing", styles["Heading2"]))
            try:
                import json
                prices = json.loads(pricing_data)
                if isinstance(prices, list) and prices:
                    price_data = [["Item", "Description", "Qty", "Unit Price", "Total"]]
                    total = 0.0
                    for item in prices:
                        qty = float(item.get("quantity", 1))
                        unit = float(item.get("unit_price", 0))
                        line_total = qty * unit
                        total += line_total
                        price_data.append([
                            item.get("item", "N/A"),
                            item.get("description", ""),
                            str(int(qty)),
                            f"${unit:,.2f}",
                            f"${line_total:,.2f}",
                        ])
                    price_data.append(["", "", "", "TOTAL:", f"${total:,.2f}"])
                    price_table = Table(price_data, colWidths=[1.2 * inch, 2.6 * inch, 0.6 * inch, 1.2 * inch, 1.2 * inch])
                    price_table.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a365d")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("ALIGN", (1, 1), (1, -2), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTNAME", (-2, -1), (-1, -1), "Helvetica-Bold"),
                        ("BACKGROUND", (-2, -1), (-1, -1), colors.HexColor("#e2e8f0")),
                    ]))
                    story.append(price_table)
            except Exception:
                story.append(Paragraph("(Pricing data could not be parsed)", body_style))

        # Terms
        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph("Terms & Conditions", styles["Heading3"]))
        terms = (
            "Payment terms: Net 30 days. "
            "All prices are in USD unless otherwise specified. "
            "This proposal is valid for 30 days from the date of issue."
        )
        story.append(Paragraph(terms, body_style))
        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph("Authorized Signature: _________________________", body_style))

        doc.build(story)
        return f"PDF generated: {file_path}"

    # Markdown fallback
    file_name = f"proposal_{safe_client}_{enquiry_id[:8]}_{timestamp}.md"
    file_path = Path("media/proposals") / file_name
    md_content = f"""# Proposal

**Client:** {client_name}
**Enquiry ID:** {enquiry_id}
**Date:** {datetime.now(timezone.utc).strftime("%Y-%m-%d")}
**Prepared by:** Aries AI Presales Consultant

---

{content}

---

*This proposal was generated by the Aries AI Presales Consultant.*
"""
    file_path.write_text(md_content, encoding="utf-8")
    return f"Markdown proposal generated: {file_path}"


@doc_output_mcp.tool()
async def generate_quote_file(
    enquiry_id: str,
    pricing_data: str,
    output_format: str = "xlsx",
) -> str:
    """Generate a quote spreadsheet (Excel or CSV).

    Args:
        enquiry_id: UUID of the enquiry.
        pricing_data: JSON array of line items [{item, description, quantity, unit_price}].
        output_format: "xlsx" or "csv".

    Returns:
        Path to the generated file.
    """
    import json

    safe_id = enquiry_id.replace("-", "_")[:20]
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    try:
        items = json.loads(pricing_data) if pricing_data else []
    except json.JSONDecodeError:
        items = []

    if not isinstance(items, list):
        items = []

    # Build rows
    rows = []
    total = 0.0
    for item in items:
        qty = float(item.get("quantity", 1))
        unit = float(item.get("unit_price", 0))
        line_total = qty * unit
        total += line_total
        rows.append({
            "Item": item.get("item", "N/A"),
            "Description": item.get("description", ""),
            "Quantity": int(qty),
            "Unit Price": f"${unit:,.2f}",
            "Total": f"${line_total:,.2f}",
        })

    rows.append({"Item": "", "Description": "", "Quantity": "", "Unit Price": "TOTAL", "Total": f"${total:,.2f}"})

    if output_format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            file_name = f"quote_{safe_id}_{timestamp}.xlsx"
            file_path = Path("media/quotes") / file_name

            wb = Workbook()
            ws = wb.active
            ws.title = "Quote"

            # Header
            headers = ["Item", "Description", "Quantity", "Unit Price", "Total"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row in rows:
                ws.append([row.get(h, "") for h in headers])

            # Style total row
            total_row = len(rows) + 1
            for cell in ws[total_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")

            # Adjust column widths
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 40
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 15

            wb.save(str(file_path))
            return f"Excel quote generated: {file_path}"
        except ImportError:
            output_format = "csv"

    # CSV fallback
    file_name = f"quote_{safe_id}_{timestamp}.csv"
    file_path = Path("media/quotes") / file_name
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Item", "Description", "Quantity", "Unit Price", "Total"])
        writer.writeheader()
        writer.writerows(rows)
    return f"CSV quote generated: {file_path}"


@doc_output_mcp.tool()
async def generate_internal_summary(
    enquiry_id: str,
    summary: str,
    outcome_status: str = "pending",
    metadata: str = "",
) -> str:
    """Generate an internal summary document (markdown).

    Args:
        enquiry_id: UUID of the enquiry.
        summary: Summary text content.
        outcome_status: e.g. "won", "lost", "pending", "negotiating".
        metadata: Optional JSON string with extra fields.

    Returns:
        Path to the generated markdown file.
    """
    import json

    safe_id = enquiry_id.replace("-", "_")[:20]
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    file_name = f"summary_{safe_id}_{timestamp}.md"
    file_path = Path("media/summaries") / file_name

    try:
        meta = json.loads(metadata) if metadata else {}
    except json.JSONDecodeError:
        meta = {}

    md_content = f"""---
type: internal-summary
enquiry_id: {enquiry_id}
status: {outcome_status}
generated_at: {datetime.now(timezone.utc).isoformat()}
---

# Internal Summary — Enquiry {enquiry_id}

**Outcome Status:** {outcome_status}
**Generated:** {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")}

---

{summary}

---

## Metadata

"""
    for key, value in meta.items():
        md_content += f"- **{key}:** {value}\n"

    md_content += "\n---\n\n*Generated by Aries AI Presales Consultant*\n"

    file_path.write_text(md_content, encoding="utf-8")
    return f"Internal summary generated: {file_path}"


@doc_output_mcp.tool()
async def generate_document(
    document_type: str,
    enquiry_id: str,
    content: str,
    client_name: str = "",
    pricing_data: str = "",
    output_format: str = "auto",
) -> str:
    """Convenience tool — routes to the correct generator based on document_type.

    Args:
        document_type: "proposal", "quote", or "summary".
        enquiry_id: UUID of the enquiry.
        content: Body text.
        client_name: Client name (for proposals).
        pricing_data: JSON pricing data (for proposals/quotes).
        output_format: "auto" picks the best format.

    Returns:
        Path to generated file.
    """
    if document_type == "proposal":
        fmt = output_format if output_format != "auto" else "pdf"
        return await generate_proposal_pdf(enquiry_id, content, client_name, pricing_data, fmt)
    elif document_type == "quote":
        fmt = output_format if output_format != "auto" else "xlsx"
        return await generate_quote_file(enquiry_id, pricing_data, fmt)
    elif document_type == "summary":
        return await generate_internal_summary(enquiry_id, content)
    else:
        return f"Unknown document_type: {document_type}. Use proposal, quote, or summary."
